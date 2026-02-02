#!/usr/bin/env python3
"""
Airtable to Excel Daily Export
Exports a specific Airtable view to an Excel file with formatting
"""

import os
import requests
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# Configuration
AIRTABLE_TOKEN = os.environ.get('AIRTABLE_TOKEN')
BASE_ID = 'appgBl5EHB3qFtOPl'
TABLE_ID = 'tbliD4adKzRMJJIXB'
VIEW_NAME = 'Indecomm Funded File'

# Email configuration
RECIPIENT_EMAIL = 'owen.sheehan@multiplymortgage.com'

# Mapping from Airtable columns to template columns
# Format: 'Template Column Name': 'Airtable Column Name'
COLUMN_MAPPING = {
    'Loan Number': 'Loan Number',
    'Borrower Name': 'Borrower Name',
    'Loan Type': 'Mortgage Type',
    'Loan Program': None,
    'Loan Purpose': None,
    'Loan Amount': 'Loan Size',
    'Lien Position': None,
    'Closing Date': 'Funded Date',
    'Funding Funds Sent Date': None,
    'Last Finished Milestone': None,
    'Subject Property Address': None,
    'Subject Property City': None,
    'Subject Property Zip': None,
    'Subject Property County': None,
    'Subject Property State': None,
    'Warehouse Co Name': None,
    'Mers Min #': None,
    'Closing Package Received Date': None,
    'Incoming Closing Package Tracking Number': None,
    'Note Tracking Number': None,
    'Trailing Doc Pulled': None,
    'Check Number': None,
    'Check Amount': None,
    'Check Tracking Number': None,
    'Comments': None,
    'Check In Date': None,
    'NOTE SHIPMENT DATE': None,
    'Package/Note Follow-up Needed?': None,
    'Package/Note Follow-up Date': None,
    'Follow-up Needed?': None,
    'Follow-up Date': None,
    'Note Received Date': None,
    'Note Received Tracking Number': None
}

def get_airtable_records():
    """Gather all records from the Airtable view"""
    url = f'https://api.airtable.com/v0/{BASE_ID}/{TABLE_ID}'
    headers = {
        'Authorization': f'Bearer {AIRTABLE_TOKEN}',
        'Content-Type': 'application/json'
    }
    
    params = {
        'view': VIEW_NAME
    }
    
    all_records = []
    offset = None
    
    print("📥 Fetching records from Airtable...")
    
    while True:
        if offset:
            params['offset'] = offset
            
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        
        data = response.json()
        records = data.get('records', [])
        all_records.extend(records)
        
        offset = data.get('offset')
        if not offset:
            break
    
    print(f"✅ Fetched {len(all_records)} records")
    return all_records

def records_to_dataframe(records):
    """Convert Airtable records to a DataFrame matching the template format"""
    if not records:
        return pd.DataFrame()
    
    # Extract fields from each record
    data = [record['fields'] for record in records]
    airtable_df = pd.DataFrame(data)
    
    # Create a new DataFrame with template columns
    template_df = pd.DataFrame()
    
    for template_col, airtable_col in COLUMN_MAPPING.items():
        if airtable_col and airtable_col in airtable_df.columns:
            # Map the Airtable column to the template column
            template_df[template_col] = airtable_df[airtable_col]
        else:
            # Create empty column
            template_df[template_col] = ''
    
    print(f"✅ Mapped {len(template_df)} rows to template format")
    
    # Log any Airtable columns that weren't mapped
    mapped_airtable_cols = [col for col in COLUMN_MAPPING.values() if col]
    unmapped = set(airtable_df.columns) - set(mapped_airtable_cols)
    if unmapped:
        print(f"ℹ️  Note: These Airtable columns were not mapped: {', '.join(unmapped)}")
    
    return template_df

def format_excel(filename):
    """Apply nice formatting to the Excel file"""
    wb = load_workbook(filename)
    ws = wb.active
    
    # Style the header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    wb.save(filename)

def send_email(filename):
    """Send the Excel file via email using Gmail SMTP"""
    smtp_server = os.environ.get('SMTP_SERVER', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', '587'))
    sender_email = os.environ.get('SENDER_EMAIL')
    sender_password = os.environ.get('SENDER_PASSWORD')
    
    if not sender_email or not sender_password:
        print("⚠️  Email credentials not configured. Skipping email send.")
        print("   Set SENDER_EMAIL and SENDER_PASSWORD in GitHub secrets to enable email.")
        return False
    
    # Create message
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = RECIPIENT_EMAIL
    msg['Subject'] = f'Daily Airtable Export - {datetime.now().strftime("%B %d, %Y")}'
    
    # Email body
    body = f"""
Hello,

Your daily Airtable export is attached.

Export Date: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}
File: {filename}

This is an automated message from your GitHub Actions workflow.
"""
    
    msg.attach(MIMEText(body, 'plain'))
    
    # Attach the Excel file
    try:
        with open(filename, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)
    except Exception as e:
        print(f"❌ Failed to attach file: {e}")
        return False
    
    # Send email
    try:
        print(f"📧 Sending email to {RECIPIENT_EMAIL}...")
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        print(f"✅ Email sent successfully to {RECIPIENT_EMAIL}")
        return True
    except Exception as e:
        print(f"❌ Failed to send email: {e}")
        return False

def main():
    """Main execution function"""
    if not AIRTABLE_TOKEN:
        raise ValueError("❌ AIRTABLE_TOKEN environment variable not set")
    
    # Fetch records
    records = get_airtable_records()
    
    if not records:
        print("⚠️  No records found in view")
        return
    
    # Convert to DataFrame
    df = records_to_dataframe(records)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d')
    filename = f'airtable_export_{timestamp}.xlsx'
    
    # Export to Excel
    print(f"💾 Exporting to {filename}...")
    df.to_excel(filename, index=False, sheet_name='Export')
    
    # Apply formatting
    print("🎨 Applying formatting...")
    format_excel(filename)
    
    print(f"✅ Export complete! File saved as: {filename}")
    print(f"📊 Exported {len(df)} rows and {len(df.columns)} columns")
    
    # Send email
    send_email(filename)

if __name__ == '__main__':
    main()
